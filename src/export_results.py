import json
import csv
import xml.etree.ElementTree as ET
from openpyxl import Workbook


class ResultExporter:
    def __init__(self, investigation_result):
        self.result = investigation_result

    def export_json(self, filename):
        with open(filename, 'w') as f:
            json.dump(self.result, f, indent=2)

    def export_csv(self, filename):
        with open(filename, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Incident ID', 'Timestamp', 'Description', 'Severity'])
            writer.writerow([
                self.result['incident']['id'],
                self.result['incident']['timestamp'],
                self.result['incident']['description'],
                self.result['incident'].get('severity', 'N/A')
            ])
            writer.writerow(['Anomaly Description', 'Confidence Score', 'Implications', 'Recommended Actions'])
            for anomaly in self.result['anomalies']:
                writer.writerow([
                    anomaly['description'],
                    anomaly['confidence_score'],
                    anomaly['implications'],
                    anomaly['recommended_actions']
                ])

    def export_xml(self, filename):
        root = ET.Element("investigation_result")
        incident = ET.SubElement(root, "incident")
        ET.SubElement(incident, "id").text = self.result['incident']['id']
        ET.SubElement(incident, "timestamp").text = self.result['incident']['timestamp']
        ET.SubElement(incident, "description").text = self.result['incident']['description']
        ET.SubElement(incident, "severity").text = str(self.result['incident'].get('severity', 'N/A'))

        understanding = ET.SubElement(root, "understanding")
        ET.SubElement(understanding, "analysis").text = self.result['understanding']['analysis']

        anomalies = ET.SubElement(root, "anomalies")
        for anomaly in self.result['anomalies']:
            anomaly_elem = ET.SubElement(anomalies, "anomaly")
            ET.SubElement(anomaly_elem, "description").text = anomaly['description']
            ET.SubElement(anomaly_elem, "confidence_score").text = str(anomaly['confidence_score'])
            ET.SubElement(anomaly_elem, "implications").text = anomaly['implications']
            ET.SubElement(anomaly_elem, "recommended_actions").text = anomaly['recommended_actions']

        tree = ET.ElementTree(root)
        tree.write(filename)

    def export_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Investigation Result"

        ws['A1'] = 'Incident ID'
        ws['B1'] = 'Timestamp'
        ws['C1'] = 'Description'
        ws['D1'] = 'Severity'

        ws['A2'] = self.result['incident']['id']
        ws['B2'] = self.result['incident']['timestamp']
        ws['C2'] = self.result['incident']['description']
        ws['D2'] = self.result['incident'].get('severity', 'N/A')

        ws['A4'] = 'Incident Understanding'
        ws['B4'] = self.result['understanding']['analysis']

        ws['A6'] = 'Anomaly Description'
        ws['B6'] = 'Confidence Score'
        ws['C6'] = 'Implications'
        ws['D6'] = 'Recommended Actions'

        for i, anomaly in enumerate(self.result['anomalies'], start=7):
            ws[f'A{i}'] = anomaly['description']
            ws[f'B{i}'] = anomaly['confidence_score']
            ws[f'C{i}'] = anomaly['implications']
            ws[f'D{i}'] = anomaly['recommended_actions']

        wb.save(filename)


# Example usage
if __name__ == "__main__":
    investigation_result = {
        "incident": {
            "id": "INC-12345",
            "timestamp": "2023-07-06T10:30:00",
            "description": "Unusual login activity detected",
            "severity": 8
        },
        "understanding": {
            "analysis": "Potential unauthorized access detected with multiple failed login attempts followed by a successful login from an unrecognized IP address."
        },
        "anomalies": [
            {
                "description": "Multiple failed login attempts from various IP addresses",
                "confidence_score": 0.95,
                "implications": "Possible brute force attack attempt",
                "recommended_actions": "Implement IP-based rate limiting and notify the account owner"
            },
            {
                "description": "Successful login from an unrecognized IP address after failed attempts",
                "confidence_score": 0.85,
                "implications": "Potential account compromise",
                "recommended_actions": "Force password reset and enable two-factor authentication"
            }
        ]
    }

    exporter = ResultExporter(investigation_result)
    exporter.export_json("result.json")
    exporter.export_csv("result.csv")
    exporter.export_xml("result.xml")
    exporter.export_excel("result.xlsx")

    print("Export completed.")
